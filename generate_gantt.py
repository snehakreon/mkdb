import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

wb = openpyxl.Workbook()

# ============================================================
# SHEET 1: STATUS SUMMARY
# ============================================================
ws1 = wb.active
ws1.title = "Project Status"

# Styles
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
section_font = Font(bold=True, size=11, color="2F5496")
section_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
done_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
partial_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
missing_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
done_font = Font(color="006100")
partial_font = Font(color="9C6500")
missing_font = Font(color="9C0006")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# Title
ws1.merge_cells("A1:D1")
ws1["A1"] = "MATERIAL KING - Project Status Report"
ws1["A1"].font = Font(bold=True, size=16, color="2F5496")

elapsed_weeks = (date(2026, 3, 21) - date(2026, 1, 22)).days / 7
remaining_weeks = (date(2026, 4, 30) - date(2026, 3, 21)).days / 7

ws1.merge_cells("A2:D2")
ws1["A2"] = f"Generated: {date.today().strftime('%d %b %Y')}  |  Team: 6-7 devs  |  Start: 22 Jan 2026  |  Deadline: 30 Apr 2026  |  {remaining_weeks:.1f} weeks remaining"
ws1["A2"].font = Font(size=10, color="666666")

# Headers
row = 4
for col, val in enumerate(["Module / Feature", "Area", "Status", "Notes"], 1):
    cell = ws1.cell(row=row, column=col, value=val)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

# Data — UPDATED TO REFLECT ACTUAL STATUS AS OF 21 MAR 2026
data = [
    # BACKEND - COMPLETED (Jan 22 - Mar 21)
    ("BACKEND - COMPLETED (Jan 22 - Mar 21)", "", "", ""),
    ("Auth Module (register, login, JWT, refresh, logout)", "Backend", "DONE", "Full JWT auth with role middleware, account locking after 5 failed attempts"),
    ("Products CRUD API", "Backend", "DONE", "List, create, update, delete + active endpoint with filters"),
    ("Brands CRUD API", "Backend", "DONE", "List, create, update, delete"),
    ("Categories CRUD API", "Backend", "DONE", "List, create, update, delete + parent/child hierarchy + reorganize endpoint"),
    ("Vendors CRUD API", "Backend", "DONE", "List, create, update, delete"),
    ("Zones CRUD API (with pincodes)", "Backend", "DONE", "Zone-vendor assignments, pincode mapping"),
    ("Buyers CRUD API (with projects)", "Backend", "DONE", "Buyer + project management"),
    ("Dealers CRUD API", "Backend", "DONE", "With credit limit, approval status"),
    ("Orders CRUD API (with items)", "Backend", "DONE", "Order creation with transaction, line items, cart auto-clear"),
    ("Cart API (server-side + guest sync)", "Backend", "DONE", "Add/update/remove/clear + sync on login"),
    ("Wishlist API", "Backend", "DONE", "Add/remove/list user wishlist"),
    ("Coupon API", "Backend", "DONE", "Validate/apply coupons, percentage/fixed, usage limits"),
    ("Admin Users API", "Backend", "DONE", "CRUD for admin user management with role assignment"),
    ("Address API", "Backend", "DONE", "CRUD for user addresses with default address support"),
    ("DB Schema (15+ tables) + Seed Script", "Backend", "DONE", "Full schema with migration patches"),
    ("Validation & Error Middleware", "Backend", "DONE", "Request validation middleware"),
    ("Role-based Access Control", "Backend", "DONE", "Admin, dealer, buyer, vendor roles"),
    ("Server-side Pagination (all endpoints)", "Backend", "DONE", "page/offset/total across all list endpoints — completed Mar 17"),
    ("File Upload Middleware (Multer)", "Backend", "DONE", "Multer middleware for Tech Data Sheet PDF uploads — completed Mar 17"),
    ("Order Workflow State Machine", "Backend", "DONE", "Transition validation (pending→confirmed→dispatched→delivered) — completed Mar 17"),
    ("Inventory Decrement on Order + Alerts", "Backend", "DONE", "Auto-decrement stock on order, low-stock validation — completed Mar 17"),

    # FRONTEND - FOUNDATION (COMPLETED)
    ("FRONTEND - FOUNDATION (COMPLETED)", "", "", ""),
    ("Vite + React 19 + TypeScript", "Frontend", "DONE", "Project scaffolded and running"),
    ("Tailwind CSS Config (custom MK theme)", "Frontend", "DONE", "mk-red, mk-gray palette, Montserrat font"),
    ("BrowserRouter + Route Definitions", "Frontend", "DONE", "Full route tree: public, admin, account sections"),
    ("React Query (QueryClientProvider)", "Frontend", "DONE", "Configured with retry:1, staleTime:30s"),
    ("Axios API Client + Token Refresh", "Frontend", "DONE", "Auto-refresh on 401, concurrent request queue"),
    ("AuthContext + useAuth hook", "Frontend", "DONE", "Login/logout/updateUser/isAdmin, auto-load on mount"),
    ("Protected Routes / Route Guards", "Frontend", "DONE", "ProtectedRoute component with loading spinner"),
    ("API Service Layer (13 modules)", "Frontend", "DONE", "product, category, brand, vendor, zone, dealer, buyer, order, auth, cart, coupon, wishlist, address, adminUser"),
    ("Login Page", "Frontend", "DONE", "Email/password login with validation"),
    ("Register Page", "Frontend", "DONE", "User registration with role selection"),

    # FRONTEND - ADMIN CRUD MODULES (COMPLETED)
    ("FRONTEND - ADMIN CRUD MODULES (COMPLETED)", "", "", ""),
    ("Admin Dashboard", "Frontend", "DONE", "Stats cards (orders, revenue, products, buyers, dealers) + recent orders table"),
    ("Products Module (list + forms)", "Frontend", "DONE", "Full CRUD with 4-tab form (basic, dimensions, attributes, packaging)"),
    ("Brands Module (list + forms)", "Frontend", "DONE", "CRUD with DataTable + FormModal"),
    ("Categories Module (list + forms)", "Frontend", "DONE", "CRUD with parent/child hierarchy"),
    ("Vendors Module (list + forms)", "Frontend", "DONE", "CRUD with DataTable + FormModal"),
    ("Zones Module (list + forms)", "Frontend", "DONE", "CRUD with DataTable + FormModal"),
    ("Buyers Module (list + forms)", "Frontend", "DONE", "CRUD with DataTable + FormModal"),
    ("Dealers Module (list + forms)", "Frontend", "DONE", "CRUD with DataTable + FormModal"),
    ("Orders Module (list + detail)", "Frontend", "DONE", "Admin order management with status updates"),
    ("Coupons Module (list + forms)", "Frontend", "DONE", "Full CRUD in both React frontend and legacy admin panel — completed Mar 21"),
    ("Admin Users Module (list + forms)", "Frontend", "DONE", "Full CRUD in both React frontend and legacy admin panel — completed Mar 21"),
    ("Reusable UI Components", "Frontend", "DONE", "DataTable, FormField, FormModal, StatusBadge"),
    ("Products File Upload UI", "Frontend", "DONE", "Tech Data Sheet PDF upload in product forms — completed Mar 17"),

    # FRONTEND - STOREFRONT (COMPLETED)
    ("FRONTEND - STOREFRONT PAGES (COMPLETED)", "", "", ""),
    ("Public Header + Nav with Category Hover Dropdowns", "Frontend", "DONE", "7 merged categories, hover subcategory menus, linked to products"),
    ("Public Footer", "Frontend", "DONE", "Footer with correct office address"),
    ("Home Page", "Frontend", "DONE", "Hero section + featured products"),
    ("Products List Page", "Frontend", "DONE", "Filters (category, brand, price range, search, sort)"),
    ("Product Detail Page", "Frontend", "DONE", "Product info tabs, wishlist, add to cart"),
    ("Categories Page", "Frontend", "DONE", "Category grid with subcategory tags"),
    ("Cart Page", "Frontend", "DONE", "Cart items, coupon apply, tax/shipping calc"),
    ("Checkout Page", "Frontend", "DONE", "Address selection, payment method, order placement"),
    ("Order Confirmation Page", "Frontend", "DONE", "Post-checkout confirmation (protected)"),
    ("About Page", "Frontend", "DONE", "Static about us page"),
    ("Contact Page", "Frontend", "DONE", "Contact form with office address"),

    # FRONTEND - BUYER ACCOUNT (COMPLETED)
    ("FRONTEND - BUYER ACCOUNT (COMPLETED)", "", "", ""),
    ("My Account Page", "Frontend", "DONE", "Profile management"),
    ("My Orders Page", "Frontend", "DONE", "Order history with detail view"),
    ("Addresses Page", "Frontend", "DONE", "Saved addresses CRUD with default"),
    ("Wishlist Page", "Frontend", "DONE", "Wishlist management"),

    # BACKEND - STILL PENDING
    ("BACKEND - PENDING (must complete by Apr 30)", "", "", ""),
    ("Payment Integration (Razorpay)", "Backend", "MISSING", "No payment gateway code. Only COD supported"),
    ("Invoice Generation (PDF)", "Backend", "MISSING", "No PDF generation library or templates"),
    ("Pricing Tiers / Discount Structures", "Backend", "PARTIAL", "Dealer credit + coupon system exists; tier-based pricing logic not implemented"),
    ("Delivery Tracking", "Backend", "PARTIAL", "Basic date fields; no tracking numbers or logistics API"),
    ("Reporting / Analytics APIs", "Backend", "MISSING", "No sales reports, trend analytics, export APIs"),
    ("Notifications (Email/SMS)", "Backend", "MISSING", "No nodemailer/twilio integration"),

    # FRONTEND - STILL PENDING
    ("FRONTEND - PENDING (must complete by Apr 30)", "", "", ""),
    ("Inventory Module UI", "Frontend", "MISSING", "Stock levels, alerts, reorder points dashboard"),
    ("Payment Module UI", "Frontend", "MISSING", "Payment processing screens (Razorpay integration)"),
    ("Invoice Module UI", "Frontend", "MISSING", "View/download invoices as PDF"),
    ("Delivery Tracking UI", "Frontend", "MISSING", "Track order delivery status with timeline"),
    ("Reporting / Dashboard Charts", "Frontend", "MISSING", "Sales summary, top products, dealer analytics with charts"),
    ("Role-based Dashboards (dealer, buyer, vendor)", "Frontend", "MISSING", "Different dashboards per role"),
    ("Responsive Design / Mobile", "Frontend", "PARTIAL", "Desktop works; mobile nav exists but needs polish"),
    ("Admin Dashboard Charts", "Frontend", "PARTIAL", "Stats cards done; no charts/graphs, no date-range analytics"),
]

row = 5
for item in data:
    name, area, status, notes = item
    # Section header
    if area == "" and status == "":
        ws1.merge_cells(f"A{row}:D{row}")
        cell = ws1.cell(row=row, column=1, value=name)
        cell.font = section_font
        cell.fill = section_fill
        cell.border = thin_border
        for c in range(2, 5):
            ws1.cell(row=row, column=c).fill = section_fill
            ws1.cell(row=row, column=c).border = thin_border
    else:
        ws1.cell(row=row, column=1, value=name).border = thin_border
        ws1.cell(row=row, column=2, value=area).border = thin_border
        status_cell = ws1.cell(row=row, column=3, value=status)
        status_cell.border = thin_border
        status_cell.alignment = Alignment(horizontal="center")
        ws1.cell(row=row, column=4, value=notes).border = thin_border

        if status == "DONE":
            status_cell.fill = done_fill
            status_cell.font = done_font
        elif status == "PARTIAL":
            status_cell.fill = partial_fill
            status_cell.font = partial_font
        elif status == "MISSING":
            status_cell.fill = missing_fill
            status_cell.font = missing_font
    row += 1

# Column widths
ws1.column_dimensions["A"].width = 55
ws1.column_dimensions["B"].width = 12
ws1.column_dimensions["C"].width = 12
ws1.column_dimensions["D"].width = 70

# ============================================================
# SHEET 2: GANTT CHART
# ============================================================
ws2 = wb.create_sheet("Gantt Chart")

# Full project: Jan 22 to Apr 30
project_start = date(2026, 1, 19)  # Monday of week containing Jan 22
today_marker = date(2026, 3, 21)
deadline = date(2026, 4, 30)
num_weeks = 15  # Jan 19 to Apr 27 (15 weeks)

# Calculate current week index (0-based)
today_week_index = (today_marker - project_start).days // 7  # = 8 (W9, week of Mar 16-22)

gantt_tasks = [
    # COMPLETED PHASE (Weeks 0-7, Jan 22 - Mar 10)
    ("COMPLETED: Backend CRUD & Setup (Jan 22 - Mar 10)", "", 0, 0, "", "done"),
    ("Project Setup (Vite, Express, DB, TypeScript)", "Full Stack", 0, 2, "All Devs", "done"),
    ("Auth Module (JWT, register, login, roles)", "Backend", 1, 2, "Dev 3, Dev 4", "done"),
    ("Products + Brands + Categories CRUD APIs", "Backend", 2, 2, "Dev 1, Dev 2", "done"),
    ("Vendors + Zones CRUD APIs", "Backend", 3, 2, "Dev 3, Dev 5", "done"),
    ("Buyers + Dealers + Orders CRUD APIs", "Backend", 4, 3, "Dev 4, Dev 6, Dev 7", "done"),
    ("Cart + Wishlist + Coupon + Address APIs", "Backend", 5, 2, "Dev 3, Dev 7", "done"),
    ("DB Schema (15+ tables) + Seed + Middleware", "Backend", 5, 2, "Dev 3", "done"),
    ("Frontend Scaffolding (Axios, AuthContext, Header)", "Frontend", 5, 2, "Dev 1", "done"),

    # PHASE 1: Foundation — COMPLETED (Mar 10-23)
    ("PHASE 1: FOUNDATION (Mar 10-23) \u2705 COMPLETED", "", 0, 0, "", ""),
    ("Fix Tailwind + BrowserRouter + React Query setup", "Frontend", 7, 1, "Dev 1", "done"),
    ("Auth Pages (Login + Register + Token Refresh)", "Frontend", 7, 1, "Dev 2", "done"),
    ("Protected Routes + useAuth hook", "Frontend", 7, 1, "Dev 1", "done"),
    ("API Service Layer (13 modules)", "Frontend", 7, 1, "Dev 1, Dev 2", "done"),
    ("Search & Filtering (product backend)", "Backend", 7, 1, "Dev 3", "done"),
    ("Storefront: Home, Products, Detail, Cart, Checkout", "Frontend", 7, 2, "Dev 1, Dev 2, Dev 5", "done"),
    ("Buyer Account: Orders, Addresses, Wishlist", "Frontend", 7, 2, "Dev 2, Dev 6", "done"),
    ("Category Reorganization + Nav Hover Dropdowns", "Frontend", 8, 1, "Dev 1", "done"),
    ("Server-side Pagination (all endpoints)", "Backend", 8, 1, "Dev 3", "done"),
    ("File Upload Middleware (Multer)", "Backend", 8, 1, "Dev 4", "done"),
    ("Order Workflow State Machine", "Backend", 8, 1, "Dev 7", "done"),
    ("Inventory Decrement on Order + Alerts", "Backend", 8, 1, "Dev 4", "done"),

    # PHASE 2: DONE EARLY + Remaining backend (Mar 24 - Apr 6)
    ("PHASE 2: CORE MODULES (Mar 24 - Apr 6) \u2014 Frontend DONE, Backend in progress", "", 0, 0, "", ""),
    ("Products Module UI (list + forms + 4 tabs)", "Frontend", 7, 2, "Dev 1", "done"),
    ("Brands + Categories Module UI", "Frontend", 7, 1, "Dev 2", "done"),
    ("Vendors Module UI", "Frontend", 7, 1, "Dev 5", "done"),
    ("Zones Module UI", "Frontend", 7, 1, "Dev 5", "done"),
    ("Dealers Module UI", "Frontend", 7, 2, "Dev 6", "done"),
    ("Buyers Module UI", "Frontend", 7, 2, "Dev 2", "done"),
    ("Orders Module UI (admin)", "Frontend", 7, 2, "Dev 1, Dev 2", "done"),
    ("Admin Dashboard (stats + recent orders)", "Frontend", 7, 2, "Dev 2, Dev 5", "done"),
    ("Coupons Module (both frontends)", "Frontend", 8, 2, "Dev 1", "done"),
    ("Admin Users Module (both frontends)", "Frontend", 8, 2, "Dev 1", "done"),
    ("Products File Upload UI (Tech Data Sheet)", "Frontend", 8, 1, "Dev 1", "done"),
    ("Pricing Tiers / Discount Backend", "Backend", 9, 1, "Dev 4", "pending"),
    ("Payment Integration Backend (Razorpay)", "Backend", 9, 2, "Dev 7", "pending"),
    ("Invoice Generation Backend (PDF)", "Backend", 10, 1, "Dev 3", "pending"),

    # PHASE 3: Orders, Business Logic & Dashboard (Apr 7-20)
    ("PHASE 3: BUSINESS LOGIC & DASHBOARD (Apr 7 - Apr 20)", "", 0, 0, "", ""),
    ("Inventory Module UI (stock levels, alerts)", "Frontend", 11, 1, "Dev 5", "pending"),
    ("Payment Module UI (Razorpay frontend)", "Frontend", 11, 1, "Dev 6", "pending"),
    ("Invoice Module UI (view/download PDF)", "Frontend", 12, 1, "Dev 6", "pending"),
    ("Admin Dashboard Charts (recharts/chart.js)", "Frontend", 11, 2, "Dev 2, Dev 5", "pending"),
    ("Reporting APIs (sales, top products, analytics)", "Backend", 11, 1, "Dev 3", "pending"),
    ("Delivery Tracking Backend + UI", "Full Stack", 11, 2, "Dev 7", "pending"),
    ("Notifications Backend (Email/SMS)", "Backend", 12, 1, "Dev 4", "pending"),

    # PHASE 4: Role-based, Polish & Launch (Apr 21-30)
    ("PHASE 4: POLISH & LAUNCH (Apr 21 - Apr 30) \U0001f680 DEADLINE", "", 0, 0, "", ""),
    ("Role-based Dashboards (dealer, buyer, vendor)", "Frontend", 13, 1, "Dev 1, Dev 6", "pending"),
    ("Reporting Dashboard Charts (frontend)", "Frontend", 13, 1, "Dev 2, Dev 5", "pending"),
    ("Responsive Design + Mobile Polish", "Frontend", 13, 1, "Dev 3", "pending"),
    ("Integration Testing (API + UI)", "Full Stack", 13, 1, "All Devs", "pending"),
    ("Bug Fixes & Performance Optimization", "Full Stack", 14, 1, "All Devs", "pending"),
    ("UAT + Production Deployment", "Full Stack", 14, 1, "All Devs", "pending"),
]

# Colors
completed_color = "A9D18E"   # Green for done
phase_colors = {
    "PHASE 1: FOUNDATION": "4472C4",
    "PHASE 1 REMAINING": "4472C4",
    "PHASE 2": "ED7D31",
    "PHASE 3": "70AD47",
    "PHASE 4": "FF6B6B",
    "COMPLETED": completed_color,
}

# Title
ws2.merge_cells("A1:F1")
ws2["A1"] = "MATERIAL KING - Gantt Chart (Jan 22, 2026 - Apr 30, 2026)"
ws2["A1"].font = Font(bold=True, size=16, color="2F5496")

ws2.merge_cells("A2:F2")
ws2["A2"] = f"Team: 6-7 Developers  |  Total: 14 Weeks  |  Elapsed: {elapsed_weeks:.0f} weeks  |  REMAINING: {remaining_weeks:.1f} weeks  |  Deadline: 30 Apr 2026"
ws2["A2"].font = Font(bold=True, size=10, color="CC0000")

# Header row
header_row = 4
headers = ["Task", "Area", "Team", "Start Date", "End Date", "Wks"]
for col, val in enumerate(headers, 1):
    cell = ws2.cell(row=header_row, column=col, value=val)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

# Week columns (start at col 7)
week_start_col = 7

today_col_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

for w in range(num_weeks):
    week_date = project_start + timedelta(weeks=w)
    col = week_start_col + w
    label = f"W{w+1}\n{week_date.strftime('%d %b')}"
    cell = ws2.cell(row=header_row, column=col, value=label)
    cell.font = Font(bold=True, size=8, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = thin_border
    ws2.column_dimensions[get_column_letter(col)].width = 8

    if w == today_week_index:
        cell.fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
    else:
        cell.fill = header_fill

# Data rows
row = 5
current_phase = ""
for task in gantt_tasks:
    name, area, start_week, duration, team, status = task

    # Phase header
    if area == "" and duration == 0:
        ws2.merge_cells(f"A{row}:{get_column_letter(week_start_col + num_weeks - 1)}{row}")
        cell = ws2.cell(row=row, column=1, value=name)
        cell.font = section_font
        cell.fill = section_fill
        cell.border = thin_border
        for c in range(2, week_start_col + num_weeks):
            ws2.cell(row=row, column=c).fill = section_fill
            ws2.cell(row=row, column=c).border = thin_border
        # Determine phase key
        if "COMPLETED" in name:
            current_phase = "COMPLETED"
        elif "PHASE 1" in name and "REMAINING" in name:
            current_phase = "PHASE 1 REMAINING"
        elif "PHASE 1" in name:
            current_phase = "PHASE 1: FOUNDATION"
        elif "PHASE 2" in name:
            current_phase = "PHASE 2"
        elif "PHASE 3" in name:
            current_phase = "PHASE 3"
        elif "PHASE 4" in name:
            current_phase = "PHASE 4"
        row += 1
        continue

    # Task info
    start_date = project_start + timedelta(weeks=start_week)
    end_date = start_date + timedelta(weeks=duration) - timedelta(days=1)

    ws2.cell(row=row, column=1, value=name).border = thin_border
    ws2.cell(row=row, column=2, value=area).border = thin_border
    ws2.cell(row=row, column=3, value=team).border = thin_border

    date_cell = ws2.cell(row=row, column=4, value=start_date)
    date_cell.number_format = "DD MMM"
    date_cell.border = thin_border
    date_cell.alignment = Alignment(horizontal="center")

    end_cell = ws2.cell(row=row, column=5, value=end_date)
    end_cell.number_format = "DD MMM"
    end_cell.border = thin_border
    end_cell.alignment = Alignment(horizontal="center")

    ws2.cell(row=row, column=6, value=duration).border = thin_border
    ws2.cell(row=row, column=6).alignment = Alignment(horizontal="center")

    # Mark done tasks differently
    if status == "done":
        name_cell = ws2.cell(row=row, column=1)
        name_cell.font = done_font

    # Gantt bars
    bar_color = phase_colors.get(current_phase, "4472C4")
    bar_fill = PatternFill(start_color=bar_color, end_color=bar_color, fill_type="solid")
    empty_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for w in range(num_weeks):
        col = week_start_col + w
        cell = ws2.cell(row=row, column=col)
        cell.border = thin_border
        if start_week <= w < start_week + duration:
            cell.fill = bar_fill
            if status == "done":
                cell.value = "\u2713"
                cell.font = Font(color="FFFFFF", bold=True, size=9)
                cell.alignment = Alignment(horizontal="center")
        else:
            cell.fill = empty_fill

    # Today marker column highlight
    today_cell = ws2.cell(row=row, column=week_start_col + today_week_index)
    if not (start_week <= today_week_index < start_week + duration):
        today_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    row += 1

# Add TODAY marker row
ws2.merge_cells(f"A{row}:F{row}")
ws2.cell(row=row, column=1, value=f"\u25b2 TODAY (Mar 21, 2026) \u2014 {remaining_weeks:.1f} weeks to deadline").font = Font(bold=True, size=11, color="CC0000")
ws2.cell(row=row, column=1).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
for c in range(2, week_start_col + num_weeks):
    ws2.cell(row=row, column=c).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
today_marker_cell = ws2.cell(row=row, column=week_start_col + today_week_index)
today_marker_cell.value = "\u25b2 NOW"
today_marker_cell.font = Font(bold=True, size=9, color="CC0000")
today_marker_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
today_marker_cell.alignment = Alignment(horizontal="center")

# Add legend
row += 2
ws2.cell(row=row, column=1, value="LEGEND:").font = Font(bold=True, size=10)
row += 1
legends = [
    (completed_color, "Completed (Jan 22 - Mar 10)"),
    ("4472C4", "Phase 1: Foundation (Mar 10-23) \u2014 100% DONE"),
    ("ED7D31", "Phase 2: Core CRUD (Mar 24 - Apr 6) \u2014 frontend DONE, backend: Razorpay/Invoice/Pricing pending"),
    ("70AD47", "Phase 3: Business Logic & Dashboard (Apr 7-20)"),
    ("FF6B6B", "Phase 4: Polish & Launch (Apr 21-30)"),
    ("CC0000", "TODAY marker (Mar 21)"),
]
for color, label in legends:
    ws2.cell(row=row, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    ws2.cell(row=row, column=2, value=label).font = Font(size=9)
    ws2.merge_cells(f"B{row}:D{row}")
    row += 1

# Column widths for Gantt
ws2.column_dimensions["A"].width = 55
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 18
ws2.column_dimensions["D"].width = 11
ws2.column_dimensions["E"].width = 11
ws2.column_dimensions["F"].width = 5

# ============================================================
# SHEET 3: SUMMARY STATS
# ============================================================
ws3 = wb.create_sheet("Summary")

ws3.merge_cells("A1:C1")
ws3["A1"] = "PROJECT SUMMARY"
ws3["A1"].font = Font(bold=True, size=14, color="2F5496")

ws3.merge_cells("A2:C2")
ws3["A2"] = f"Start: 22 Jan 2026  |  Today: 21 Mar 2026  |  Deadline: 30 Apr 2026"
ws3["A2"].font = Font(bold=True, size=10, color="CC0000")

stats = [
    ("", "", ""),
    ("TIMELINE", "", ""),
    ("Project Start", "22 Jan 2026", ""),
    ("Today", "21 Mar 2026", f"{elapsed_weeks:.0f} weeks elapsed"),
    ("Hard Deadline", "30 Apr 2026", f"{remaining_weeks:.1f} weeks remaining"),
    ("Total Duration", "14 weeks", "~1 week buffer if on track"),
    ("", "", ""),
    ("COMPLETION STATUS", "Count", "Details"),
    ("Backend CRUD Completed", "21 modules", "Auth, Products, Brands, Categories, Vendors, Zones, Buyers, Dealers, Orders, Cart, Wishlist, Coupon, Admin Users, Address, Schema, Seed, Validation, RBAC, Pagination, File Upload, Order State Machine, Inventory"),
    ("Backend Features Pending", "4 features", "Payments (Razorpay), Invoices (PDF), Reporting/Analytics APIs, Notifications (Email/SMS)"),
    ("Backend Features Partial", "2 features", "Pricing Tiers (coupon exists, tier logic pending), Delivery Tracking (basic fields, no logistics API)"),
    ("Frontend Completed", "54 items", "All admin CRUD (11 modules incl. Coupons & Admin Users), all storefront pages (11), buyer account (4), foundation (10), UI components (4), file upload UI, nav/footer"),
    ("Frontend Pending", "7 items", "Inventory UI, Payment UI, Invoice UI, Delivery UI, Dashboard Charts, Role-based Dashboards, Responsive polish"),
    ("", "", ""),
    ("Overall Completion", "~70%", "Phase 1 100% done. Phase 2 frontend done. Remaining: 4 backend features + 7 frontend UIs + testing/deployment."),
    ("", "", ""),
    ("RISK ASSESSMENT", "", ""),
    ("Schedule Risk", "LOW-MEDIUM", "Ahead of schedule. Phase 1 fully done. ~30% work in 5.7 weeks. Comfortable pace."),
    ("Critical Path", "Razorpay + Invoice PDF", "Payment backend needed before Payment UI; Invoice backend needed before Invoice UI"),
    ("Parallel Tracks Needed", "YES", "Backend (Razorpay, Invoice, Reporting) + Frontend (Inventory UI, Charts) simultaneously"),
    ("Recommended Action", "START PHASE 2 BACKEND", "Begin Razorpay integration and Invoice PDF generation next week"),
    ("", "", ""),
    ("PHASE PLAN (REMAINING)", "Duration", "Focus"),
    ("Phase 1", "COMPLETED \u2705", "All items done: Pagination, File Upload, Order State Machine, Inventory Decrement"),
    ("Phase 2 Remaining", "Mar 24 - Apr 6 (2 wks)", "Pricing tiers, Razorpay payment backend, Invoice PDF generation"),
    ("Phase 3: Business Logic", "Apr 7-20 (2 wks)", "Inventory UI, Payment UI, Invoice UI, Dashboard charts, Reporting APIs, Delivery tracking, Notifications"),
    ("Phase 4: Polish & Launch", "Apr 21-30 (1.5 wks)", "Role-based dashboards, responsive design, testing, UAT, deployment"),
    ("", "", ""),
    ("WHAT WAS DONE AHEAD OF SCHEDULE", "", ""),
    ("All Admin CRUD UIs (11 modules)", "Phase 2 \u2192 Done in Phase 1", "Products, Brands, Categories, Vendors, Zones, Buyers, Dealers, Orders, Dashboard, Coupons, Admin Users"),
    ("All Storefront Pages (11 pages)", "Phase 2-3 \u2192 Done in Phase 1", "Home, Products, Detail, Cart, Checkout, Confirmation, Categories, About, Contact + Header/Footer"),
    ("Buyer Account (4 pages)", "Phase 3 \u2192 Done in Phase 1", "My Account, My Orders, Addresses, Wishlist"),
    ("Cart + Wishlist + Coupon APIs", "Not in plan \u2192 Done", "Server-side cart, wishlist, coupon validation \u2014 bonus features"),
    ("Phase 1 Backend (4 items)", "Done on time (Mar 17)", "Pagination, Multer upload, Order state machine, Inventory decrement"),
    ("Coupons + Admin Users (legacy admin)", "Done Mar 21", "Full CRUD modules added to both React frontend and legacy admin panel"),
    ("", "", ""),
    ("TEAM ALLOCATION (suggested for remaining)", "", ""),
    ("Dev 1 (Frontend Lead)", "", "Inventory UI \u2192 Role-based Dashboards \u2192 Responsive polish"),
    ("Dev 2 (Frontend)", "", "Dashboard Charts \u2192 Reporting UI \u2192 Integration testing"),
    ("Dev 3 (Backend + Frontend)", "", "Invoice PDF backend \u2192 Reporting APIs \u2192 Invoice UI"),
    ("Dev 4 (Backend Lead)", "", "Pricing Tiers \u2192 Notifications (Email/SMS)"),
    ("Dev 5 (Frontend)", "", "Dashboard Charts \u2192 Responsive design"),
    ("Dev 6 (Frontend)", "", "Payment UI \u2192 Invoice UI \u2192 Role-based views"),
    ("Dev 7 (Backend)", "", "Razorpay integration \u2192 Delivery tracking backend + UI"),
]

for i, (a, b, c) in enumerate(stats, 4):
    ws3.cell(row=i, column=1, value=a).border = thin_border if a else Border()
    ws3.cell(row=i, column=2, value=b).border = thin_border if a else Border()
    ws3.cell(row=i, column=3, value=c).border = thin_border if a else Border()
    if a in ("COMPLETION STATUS", "PHASE PLAN (REMAINING)", "TEAM ALLOCATION (suggested for remaining)",
             "TIMELINE", "RISK ASSESSMENT", "WHAT WAS DONE AHEAD OF SCHEDULE"):
        for col in range(1, 4):
            ws3.cell(row=i, column=col).font = header_font
            ws3.cell(row=i, column=col).fill = header_fill
    if a == "Overall Completion":
        ws3.cell(row=i, column=1).font = Font(bold=True, size=12, color="006100")
        ws3.cell(row=i, column=2).font = Font(bold=True, size=12, color="006100")
        ws3.cell(row=i, column=3).font = Font(bold=True, size=11, color="006100")
    if "MEDIUM" in str(b):
        ws3.cell(row=i, column=2).font = Font(bold=True, color="ED7D31")
        ws3.cell(row=i, column=2).fill = partial_fill

ws3.column_dimensions["A"].width = 42
ws3.column_dimensions["B"].width = 34
ws3.column_dimensions["C"].width = 80

# Save
output_path = "/home/user/materialking/MaterialKing_Gantt_Chart.xlsx"
wb.save(output_path)
print(f"Gantt chart saved to: {output_path}")
